#This script generates the triples concerning the Narrow/Broad relations between the codes, the status of the codes, the replacedBy/replaces relations for the codes that have a change of type (recoded/boundary shift/boundary change)
#The input is the first sheet of the NUTS spreadsheet. The spreadsheet must be saved as ".xlsx"

from openpyxl import*
import re
import sys

#Maximum number of lines that the spreadsheet to process contains. Please update with the correct number
max_lines = int(sys.argv[1])
#Name of the input spreadsheet to process
input_name = sys.argv[2]


#Definition of methods
def Replace(b,c,h,count):
	if re.search('recoded',h):
		print('recoded')
		rdf_file.write('<http://data.europa.eu/nuts/code/'+b+'> <http://purl.org/dc/terms/isReplacedBy> <http://data.europa.eu/nuts/code/'+c+'> .\n')
		rdf_file.write('<http://data.europa.eu/nuts/code/'+c+'> <http://purl.org/dc/terms/Replaces> <http://data.europa.eu/nuts/code/'+b+'> .\n')
		count+=1
	if re.search('boundary shift',h):
		print('boundary shift')
		rdf_file.write('<http://data.europa.eu/nuts/code/'+b+'> <http://purl.org/dc/terms/isReplacedBy> <http://data.europa.eu/nuts/code/'+c+'> .\n')
		rdf_file.write('<http://data.europa.eu/nuts/code/'+c+'> <http://purl.org/dc/terms/Replaces> <http://data.europa.eu/nuts/code/'+b+'> .\n')
		count+=1
	if re.search('boundary change',h):
		print('boundary shift')
		rdf_file.write('<http://data.europa.eu/nuts/code/'+b+'> <http://purl.org/dc/terms/isReplacedBy> <http://data.europa.eu/nuts/code/'+c+'> .\n')
		rdf_file.write('<http://data.europa.eu/nuts/code/'+c+'> <http://purl.org/dc/terms/Replaces> <http://data.europa.eu/nuts/code/'+b+'> .\n')
		count+=1
	return count

def Status(b,c):
		if b == c:
			rdf_file.write('<http://data.europa.eu/nuts/code/'+c+'> <http://www.w3.org/ns/adms#status> <http://publications.europa.eu/resource/authority/concept-status/CURRENT> .\n')
		elif b != c:
			if c == 'None':
				rdf_file.write('<http://data.europa.eu/nuts/code/'+b+'> <http://www.w3.org/ns/adms#status> <http://publications.europa.eu/resource/authority/concept-status/DEPRECATED> .\n')
			elif b == 'None':
				rdf_file.write('<http://data.europa.eu/nuts/code/'+c+'> <http://www.w3.org/ns/adms#status> <http://publications.europa.eu/resource/authority/concept-status/CURRENT> .\n')
			else:
				rdf_file.write('<http://data.europa.eu/nuts/code/'+b+'> <http://www.w3.org/ns/adms#status> <http://publications.europa.eu/resource/authority/concept-status/DEPRECATED> .\n')
				rdf_file.write('<http://data.europa.eu/nuts/code/'+c+'> <http://www.w3.org/ns/adms#status> <http://publications.europa.eu/resource/authority/concept-status/CURRENT> .\n')
		
#Open workbook
wb = Workbook()
wb = load_workbook(filename = input_name)
#Use code below if you want to be prompted for the input file
#wb = load_workbook(filename = str(input("Give name of output file: ")))  

rdf_file=open('NUTS_changes.txt','w')
#Use code below if you want to manually name the output file
#rdf_file=open(str(input("Give name of output file: ")),'w') 


#Print sheet names
print(wb.get_sheet_names())

#Choose sheet
sheet = wb['NUTS2013-NUTS2016']
#sheet = wb[str(input("Give name of sheet containing the NUTS data: "))]

count=0
count_total=0

#1. Replaces/ReplacedBy triples

#Aqcuire cell values
i=3
while i < max_lines:
#while i < sheet.max_row:
#The function above can be used instead of hardcoding the maximum rows, but seems to make the script slower. Furtheromre the function seems to calculate more rows han contain data (probably because of formatting or something in Excel.
	
	colB='B'+str(i)
	colC='C'+str(i)
	colH='H'+str(i)
	if sheet[colB].value != sheet[colC].value:
		print('\nchange detected')
		count_total+=1
		#Here call RDF function 
		count=Replace(str(sheet[colB].value),str(sheet[colC].value),str(sheet[colH].value),count)

	i+=1

print('\n')
print(count)
print(count_total)

#2. Narrow-Broad triples

i=3
while i < max_lines:
#while i < sheet.max_row:

	colB='B'+str(i)
	colI='I'+str(i)

	
	if sheet[colB].value is not None:
		if sheet[colI].value == 0:
			colB0='B'+str(i)
		elif sheet[colI].value == 1:
			colB1='B'+str(i)
			rdf_file.write('<http://data.europa.eu/nuts/code/'+sheet[colB].value+'> <http://www.w3.org/2004/02/skos/core#broader> <http://data.europa.eu/nuts/code/'+sheet[colB0].value+'> .\n')
			rdf_file.write('<http://data.europa.eu/nuts/code/'+sheet[colB0].value+'> <http://www.w3.org/2004/02/skos/core#narrower> <http://data.europa.eu/nuts/code/'+sheet[colB].value+'> .\n')
		elif sheet[colI].value == 2:
			colB2='B'+str(i)
			rdf_file.write('<http://data.europa.eu/nuts/code/'+sheet[colB].value+'> <http://www.w3.org/2004/02/skos/core#broader> <http://data.europa.eu/nuts/code/'+sheet[colB1].value+'> .\n')
			rdf_file.write('<http://data.europa.eu/nuts/code/'+sheet[colB1].value+'> <http://www.w3.org/2004/02/skos/core#narrower> <http://data.europa.eu/nuts/code/'+sheet[colB].value+'> .\n')
		elif sheet[colI].value == 3:
			colB3='B'+str(i)
			rdf_file.write('<http://data.europa.eu/nuts/code/'+sheet[colB].value+'> <http://www.w3.org/2004/02/skos/core#broader> <http://data.europa.eu/nuts/code/'+sheet[colB2].value+'> .\n')
			rdf_file.write('<http://data.europa.eu/nuts/code/'+sheet[colB2].value+'> <http://www.w3.org/2004/02/skos/core#narrower> <http://data.europa.eu/nuts/code/'+sheet[colB].value+'> .\n')

	i+=1

#3. Status triples

i=3
while i < max_lines:
	
	colB='B'+str(i)
	colC='C'+str(i)

	Status(str(sheet[colB].value),str(sheet[colC].value))

	i+=1

	#print('Completed '+str(int(i/1868*100))+'%')

print(sheet.max_row)


rdf_file.close()